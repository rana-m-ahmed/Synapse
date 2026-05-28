"""
Synapse Backend — Excel Parser
=================================
Parses Excel (.xlsx) files using openpyxl.
Handles multi-sheet workbooks by processing each sheet separately.
Uses the same row-to-sentence approach as the CSV parser for RAG-friendly output.
"""

import io
import logging

from openpyxl import load_workbook

from app.core.exceptions import DocumentProcessingError
from app.document_processing.parsers.base import BaseParser, ParsedDocument

logger = logging.getLogger("synapse.parsers.excel")


class ExcelParser(BaseParser):
    """
    Parses Excel files by extracting text from all sheets.
    Each row is converted to a readable sentence using column headers as labels.
    Sheet names are included in metadata for multi-sheet workbooks.
    """

    async def parse(self, file_content: bytes, file_name: str) -> ParsedDocument:
        """
        Parse an Excel (.xlsx) file into natural language text.

        Processes each sheet independently. The first row of each sheet is
        treated as headers. Subsequent rows are converted to:
        "Sheet: SheetName — column1: value1, column2: value2, ..."

        Args:
            file_content: Raw Excel file bytes.
            file_name: Original filename.

        Returns:
            ParsedDocument with row-sentence text from all sheets.

        Raises:
            DocumentProcessingError: If the file can't be read or is empty.
        """
        try:
            workbook = load_workbook(
                io.BytesIO(file_content),
                read_only=True,  # Memory efficient for large files
                data_only=True,  # Read cell values, not formulas
            )
        except Exception as e:
            logger.error(f"Failed to read Excel file '{file_name}': {e}")
            raise DocumentProcessingError(
                f"Could not read Excel file '{file_name}'. The file may be corrupt."
            )

        sheet_names = workbook.sheetnames
        if not sheet_names:
            raise DocumentProcessingError(
                f"Excel file '{file_name}' has no sheets."
            )

        all_sentences: list[str] = []
        total_rows = 0

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows or len(rows) < 2:
                # Need at least a header row + one data row
                continue

            # First row is headers
            headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(rows[0])]

            # Process data rows
            for row in rows[1:]:
                total_rows += 1
                parts = []
                for header, value in zip(headers, row):
                    if value is not None and str(value).strip():
                        parts.append(f"{header}: {str(value).strip()}")

                if parts:
                    # Include sheet name for multi-sheet workbooks
                    if len(sheet_names) > 1:
                        sentence = f"Sheet: {sheet_name} — " + ", ".join(parts)
                    else:
                        sentence = ", ".join(parts)
                    all_sentences.append(sentence)

        workbook.close()

        if not all_sentences:
            raise DocumentProcessingError(
                f"Excel file '{file_name}' contains no data."
            )

        full_text = "\n".join(all_sentences)

        logger.info(
            f"Parsed Excel '{file_name}': {len(sheet_names)} sheets, "
            f"{total_rows} rows, {len(full_text)} chars"
        )

        return ParsedDocument(
            text=full_text,
            metadata={
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "row_count": total_rows,
                "source_file": file_name,
            },
        )
